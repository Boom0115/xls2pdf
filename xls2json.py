#! python3
import openpyxl, pprint
import json

class Xls2json:


    def __init__(self, in_xls, out_json):
        self.title_row = 4

        self.in_xls = in_xls
        self.out_json_file = out_json
        # XLSXを開く
        self.wb = openpyxl.load_workbook(self.in_xls)


    def getLayout(self):
        '''TODO:Layoutシートからレイアウト情報を取得する'''
        ws = self.wb.get_sheet_by_name('layout1')
        layout_info = {}

        for row in ws.iter_rows():
            for cell in row:
                print(cell, ":" , cell.value, cell.internal_value)

    def export(self):
        sheet = self.wb.get_sheet_by_name('card_info')

        title_info = {}
        # タイトルを取得
        title_info['title'] = str(sheet.cell('B1').value)
        title_info['version'] = str(sheet.cell('B2').value)
        title_info['author'] = str(sheet.cell('D2').value)


        for cell in sheet[self.title_row]:
            print(cell.value)

        # 見出し行を取得
        labels = []
        for cell in sheet[self.title_row]:
            labels.append(cell.value)

        # カード情報を構築する
        card_info = []
        for row in range(self.title_row + 1, sheet.max_row):
            rec = []
            cols = sheet[row]
            for i, cell in enumerate(cols):
                rec.append((labels[i], cell.value))
            card_info.append(tuple(rec))

        # カード情報を.json ファイルに出力する
        full_info = {
            'title_info': title_info,
            'card_info': card_info,
        }

        with open(self.out_json_file, mode='w', encoding='utf-8') as f:
            json.dump(full_info, f
                      , indent=2                # 空白インデント数
                      , ensure_ascii=False      # 非アスキー文字をエスケープするか
                      )

if __name__ == '__main__':
    i_file = './in/cards.xlsx'

    o_file = './tmp/tmp_card_info.json'
    excel_to_json = Xls2json(i_file, o_file)
    #excel_to_json.getLayout()
    excel_to_json.export()